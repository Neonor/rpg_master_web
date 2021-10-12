from os import listdir
from io import BytesIO
import math as np
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

def new_xlsx():
    new_wb = Workbook()
    ws_data = new_wb.active
    ws_data.title = "Carte"
    
    ws_data["A1"] = "Active"
    ws_data["B1"] = "X"
    ws_data["C1"] = "Y"
    ws_data["D1"] = "BG color"
    ws_data["E1"] = "BG"
    ws_data["F1"] = "img"
    ws_data["G1"] = "Border color"
    ws_data["H1"] = "icon"

    ws_data["A2"] = 3
    ws_data["B2"] = 0
    ws_data["C2"] = 0
    ws_data["D2"].fill = PatternFill(fgColor="96B469", fill_type = "solid")
    ws_data["E2"] = "scrub"
    ws_data["F2"] = "big_city_walled"
    # ws_data["G2"] = "Border color"
    ws_data["H2"] = "crown"
    
    dv_bg = DataValidation(type="list", formula1='"%s"' % ",".join([enter.split(".")[0] for enter in listdir("cartemaker/imgs/bg") if ".png" in enter]), allow_blank=True)
    ws_data.add_data_validation(dv_bg)
    dv_bg.add('E2:E1048576')
    
    dv_obj = DataValidation(type="list", formula1='"%s"' % ",".join([enter.split(".")[0] for enter in listdir("cartemaker/imgs/obj") if ".png" in enter]), allow_blank=True)
    ws_data.add_data_validation(dv_obj)
    dv_obj.add('F2:F1048576')
    
    dv_ico = DataValidation(type="list", formula1='"%s"' % ",".join([enter.split(".")[0] for enter in listdir("cartemaker/imgs/icon") if ".png" in enter]), allow_blank=True)
    ws_data.add_data_validation(dv_ico)
    dv_ico.add('H2:H1048576')

    

    excel_file = BytesIO()
    new_wb.save(excel_file)
    
    new_wb.close()
    excel_file.seek(0)
    return excel_file

def decode_color(value):
    return (int(value[2:4],16),int(value[4:6],16),int(value[6:8],16))

def mask_hexa(x,y,r):
    x,y = x//2,y//2

    cos = np.cos(np.pi/6)

    p1 = x+r*cos,y+r/2
    p2 = x,y+r
    p3 = x-r*cos,y+r/2
    p4 = x-r*cos,y-r/2
    p5 = x,y-r
    p6 = x+r*cos,y-r/2

    mask = Image.new('RGBA', (x*2,y*2))
    d = ImageDraw.Draw(mask)
    d.polygon((p1,p2,p3,p4,p5,p6), fill='#000')
    return mask

class CarteMaker(object):
    def __init__(self,*_,size=None,r=200,struct_map=[]):
        self.size = size or (0,0)
        self.r = r
        self.offset = (0,0)
        
        self.struct_map = struct_map
        self.struct_map.sort(key=lambda x:x[0])
        if struct_map and not size:
            _,x,y,*_ = list(zip(*struct_map))
            self.offset = (min(x),min(y))
            self.size = (max(x)-min(x)+1,max(y)-min(y)+1)
        
        self._cos = np.cos(np.pi/6)
        
#         self.bgs_colors = {enter.split(".")[0]:self.find_color("./imgs/bg/%s" % enter) for enter in listdir("./imgs/bg") if ".png" in enter}
        self.bgs = {enter.split(".")[0]:self.make_bg_hexa("cartemaker/imgs/bg/%s" % enter) for enter in listdir("cartemaker/imgs/bg") if ".png" in enter}
        self.objs = {enter.split(".")[0]:self.make_obj("cartemaker/imgs/obj/%s" % enter) for enter in listdir("cartemaker/imgs/obj") if ".png" in enter}
        self.icons = {enter.split(".")[0]:self.make_obj("cartemaker/imgs/icon/%s" % enter) for enter in listdir("cartemaker/imgs/icon") if ".png" in enter}
        
        self.carte = self.make_carte()
        
    def make_obj(self,filename):
        im = Image.open(filename)
        x,y = im.size
        coef = 1.0
        y = int(coef * y * 2*int(self.r*self._cos)//x)
        x = int(coef * 2*int(self.r*self._cos))
        return im.resize((x,y))
        
        
    def make_bg_hexa(self,filename):
        im = Image.open(filename)
        x,y = im.size
        r = int(0.45*min([x,y]))
        try:
            out = Image.new('RGBA', (2*int(r*self._cos),r*2),(255,255,255,0))
            mask = mask_hexa(*im.size,r)
            im = im.resize(mask.size)
            out.paste(im, (-(x//2-int(r*self._cos)),-(y//2-r)), mask)
            return out.resize((2*int(self.r*self._cos),2*self.r))
        except:
            print(filename)
            return
        

    
    def make_border_hexa(self,im_name,border_size,bg_rgb,bd_rgb,obj_name,icon_name):
        im = self.bgs.get(im_name)
        obj = self.objs.get(obj_name)
        icon = self.icons.get(icon_name)
        im_size = (2*int(self.r*self._cos),2*self.r)
            
        border_size = (border_size//2)*2+(border_size%2 and 2 or 0)
        
#         bg_rgb = self.bgs_colors.get(im_name,bg_rgb)

        bg = Image.new('RGBA', im_size,(*bg_rgb,0))
        
        bg.paste(Image.new('RGBA', im_size,bd_rgb),
                 (0,0),
                 mask_hexa(*im_size,self.r))

        bg.paste(Image.new('RGBA', [im_size[0]-border_size,im_size[1]-border_size],(*bg_rgb,255)),
                 (border_size//2,border_size//2),
                 mask_hexa(im_size[0]-border_size,im_size[1]-border_size,(im_size[1]-border_size)//2))
        
        if im:
            temp = Image.new('RGBA', im_size,(0,0,0,0))
            temp.paste(im,
                       (0,0),
                       mask_hexa(*im_size,(im_size[1]-border_size)//2))
            bg.alpha_composite(temp)
        
        if obj:
            temp = Image.new('RGBA', im_size,(0,0,0,0))
            x,y = obj.size
            xh,yh = im_size
            temp.paste(obj,((xh-x)//2,(yh-y)//2),obj)
            bg.alpha_composite(temp)

        if icon:
            temp = Image.new('RGBA', im_size,(0,0,0,0))
            icon = icon.copy().resize((icon.size[0]//5,icon.size[1]//5))
            
            x,y = icon.size
            xh,yh = im_size
            temp.paste(Image.new('RGBA', icon.size,(0,0,0,255)),(xh-x-border_size+3,(yh-y)//4+border_size+3),icon)
            temp.paste(icon,(xh-x-border_size,(yh-y)//4+border_size),icon)
            bg.alpha_composite(temp)

        return bg

    def pos(self,x,y):
        x,y = x-self.offset[0],y-self.offset[1]
        out_x,out_y = 0,0
        if x:
            out_x += x*2*int(self.r*self._cos)
        if y:
            out_y += y*int(1.5*self.r)
        if y%2:
            out_x += int(self.r*self._cos)
        return out_x,out_y

    def make_carte(self):
        x,y = self.size
        out_x,out_y = 0,0
        out_y = int(y*1.5*self.r+self.r/2)
        out_x = x*int(2*self.r*self._cos)
#         if y>1:
#             out_x += int(self.r*self._cos)
        return Image.new('RGBA', (out_x,out_y),(255,255,255,0))
    
    def append_hexa(self,im_name,coord,bg_rgb=(0,0,0),bd_rgb=(0,0,0,255),border_size=10,obj=None,icon=None):
        im = self.make_border_hexa(im_name,border_size=border_size,bg_rgb=bg_rgb,bd_rgb=bd_rgb,obj_name=obj,icon_name=icon)
        self.carte.paste(im,self.pos(*coord),im)
    
    def aff_coord(self,color=(255,255,100)):
        carte = self.carte.copy()
        obj = ImageDraw.Draw(carte)
        fnt = ImageFont.truetype('./aAhaWow.ttf', 100)
        for _,x,y,*_ in self.struct_map:
            pos_x,pos_y = self.pos(x,y)
            obj.text((pos_x+3+int(self.r*self._cos)//2,pos_y+5+self.r//2), "%i/%i" % (x,y), font=fnt,fill=(0,0,0))
            obj.text((pos_x+int(self.r*self._cos)//2,pos_y+self.r//2), "%i/%i" % (x,y), font=fnt,fill=color)

        return carte

    def generer(self):
        for active,x,y,bg_rgb,bg,obj,bd_rgb,icon in self.struct_map:
            self.append_hexa(bg,(x,y),bg_rgb,bd_rgb,10,active>1 and obj or None,active>2 and icon or None)
        
    def reset(self):
        self.carte = self.make_carte()
