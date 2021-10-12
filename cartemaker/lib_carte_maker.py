from os import listdir,path,mkdir
from io import BytesIO
import math as np
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

CONV_HEAD = {
    'Active':"active",
    'X':"x",
    'Y':"y",
    'BG color':"bg_rgb",
    'BG':"bg",
    'img':"obj",
    'Border color':"bd_rgb",
    'icon':"icon",
    "img size":"obj_size"
}

DEFAULT = {
    "active":0,
    "bg":"",
    "obj":"",
    "obj_size":1
}

def new_xlsx():
    new_wb = Workbook()
    ws_data = new_wb.active
    ws_data.title = "Carte"
    
    for idx,key in enumerate(CONV_HEAD.keys()):
        ws_data.cell(1,idx+1).value = key

    ws_data["A2"] = 3
    ws_data["B2"] = 0
    ws_data["C2"] = 0
    ws_data["D2"].fill = PatternFill(fgColor="96B469", fill_type = "solid")
    ws_data["E2"] = "scrub"
    ws_data["F2"] = "big_city_walled"
    # ws_data["G2"] = "Border color"
    ws_data["H2"] = "crown"
    ws_data["I2"] = 1
    
    bg = [enter.split(".")[0] for enter in listdir("cartemaker/imgs/bg") if ".png" in enter]
    bg.sort()
    dv_bg = DataValidation(type="list", formula1='"%s"' % ",".join(bg), allow_blank=True)
    ws_data.add_data_validation(dv_bg)
    dv_bg.add('E2:E1048576')
    
    obj = [enter.split(".")[0] for enter in listdir("cartemaker/imgs/obj") if ".png" in enter]
    obj.sort()
    dv_obj = DataValidation(type="list", formula1='"%s"' % ",".join(obj), allow_blank=True)
    ws_data.add_data_validation(dv_obj)
    dv_obj.add('F2:F1048576')
    
    ico = [enter.split(".")[0] for enter in listdir("cartemaker/imgs/icon") if ".png" in enter]
    ico.sort()
    dv_ico = DataValidation(type="list", formula1='"%s"' % ",".join(ico), allow_blank=True)
    ws_data.add_data_validation(dv_ico)
    dv_ico.add('H2:H1048576')

    

    excel_file = BytesIO()
    new_wb.save(excel_file)
    
    new_wb.close()
    excel_file.seek(0)
    return excel_file

def decode_color(value):
    return (int(value[2:4],16),int(value[4:6],16),int(value[6:8],16))

def mask_hexa(x,y,r):
    x,y = x//2,y//2

    cos = np.cos(np.pi/6)

    p1 = x+r*cos,y+r/2
    p2 = x,y+r
    p3 = x-r*cos,y+r/2
    p4 = x-r*cos,y-r/2
    p5 = x,y-r
    p6 = x+r*cos,y-r/2

    mask = Image.new('RGBA', (x*2,y*2))
    d = ImageDraw.Draw(mask)
    d.polygon((p1,p2,p3,p4,p5,p6), fill='#000')
    return mask



class DataCase(object):
    def __init__(self,data):
        for enter in CONV_HEAD.values():
            setattr(self,enter,None)

        for f_head,value in data.items():
            head = CONV_HEAD.get(f_head,f_head)
            value = value.value is None and (DEFAULT.get(head) or decode_color(value.fill.fgColor.value)) or value.value
            setattr(self,head,value)


def make_carte(file_post):
    destination = BytesIO()
    for chunk in file_post.chunks():
        destination.write(chunk)
    destination.seek(0)
    wb = load_workbook(filename=destination)

    heads = [value.value for value in wb["Carte"][1]]

    struct_map = [DataCase(dict(zip(heads,enter)))
          for enter in wb["Carte"]
          if enter[0].value and enter[0].row >1]
    wb.close()
    carte = CarteMaker(struct_map=struct_map)
    carte.generer()

    img_byte_arr = BytesIO()
    carte.carte.save(img_byte_arr, format='PNG')
    img_byte_arr.seek(0)
    return img_byte_arr

IMGS_FOLDER = "cartemaker/imgs"
CACHE_FOLDER = "cartemaker/cache"

class CarteMaker(object):
    def __init__(self,*,size=None,r=200,struct_map=[]):
        self.size = size or (0,0)
        self.r = r
        self.offset = (0,0)
        
        self.struct_map = struct_map
        self.struct_map.sort(key=lambda x:x.active)
        if struct_map and not size:
            x = [enter.x for enter in struct_map]
            y = [enter.y for enter in struct_map]
            self.offset = (min(x),min(y))
            self.size = (max(x)-min(x)+1,max(y)-min(y)+1)
        
        border_size = self.r//20
        self.border_size = (border_size//2)*2+(border_size%2 and 2 or 0)
        
        self._cos = np.cos(np.pi/6)
        
        bg = [enter.bg for enter in self.struct_map]
        obj = [enter.obj for enter in self.struct_map]
        icon = [enter.icon for enter in self.struct_map]

        self.bgs = {enter.split(".")[0]:self.make_bg_hexa(enter) for enter in listdir(path.join(IMGS_FOLDER,"bg")) if ".png" in enter and enter.split(".")[0] in bg}
        self.objs = {enter.split(".")[0]:self.make_obj("obj", enter) for enter in listdir(path.join(IMGS_FOLDER,"obj")) if ".png" in enter and enter.split(".")[0] in obj}
        self.icons = {enter.split(".")[0]:self.make_obj("icon", enter) for enter in listdir(path.join(IMGS_FOLDER,"icon")) if ".png" in enter and enter.split(".")[0] in icon}
        
        self.carte = self.make_carte()
        
    def make_obj(self,folder,filename):
        if path.isfile(path.join(CACHE_FOLDER,folder,"%i_%s" % (self.r,filename))):
            return Image.open(path.join(CACHE_FOLDER,folder,"%i_%s" % (self.r,filename)))
        else:
            im = Image.open(path.join(IMGS_FOLDER,folder,filename))
            x,y = im.size
            coef = 1.0
            y = int(coef * y * 2*int(self.r*self._cos)//x)
            x = int(coef * 2*int(self.r*self._cos))
            im = im.resize((x,y))
            if not path.isdir(path.join(CACHE_FOLDER,folder)):
                mkdir(path.join(CACHE_FOLDER,folder))
            im.save(path.join(CACHE_FOLDER,folder,"%i_%s" % (self.r,filename)), format='PNG')
            return im
        
        
    def make_bg_hexa(self,filename):
        if path.isfile(path.join(CACHE_FOLDER,"bg","%i_%s" % (self.r,filename))):
            return Image.open(path.join(CACHE_FOLDER,"bg","%i_%s" % (self.r,filename)))
        else:
            im = Image.open(path.join(IMGS_FOLDER,"bg",filename))

            x,y = im.size
            r = int(0.45*min([x,y]))
            try:
                out = Image.new('RGBA', (2*int(r*self._cos),r*2),(255,255,255,0))
                mask = mask_hexa(*im.size,r)
                im = im.resize(mask.size)
                out.paste(im, (-(x//2-int(r*self._cos)),-(y//2-r)), mask)
                im = out.resize((2*int(self.r*self._cos),2*self.r))
                if not path.isdir(path.join(CACHE_FOLDER,"bg")):
                    mkdir(path.join(CACHE_FOLDER,"bg"))
                im.save(path.join(CACHE_FOLDER,"bg","%i_%s" % (self.r,filename)), format='PNG')
                return im
            except:
                print(filename)
                return

    def make_bg(self):
        """génère le fond et les bords"""        
        im_size = (2*int(self.r*self._cos),2*self.r)

        mask1 = mask_hexa(*im_size,self.r)
        mask2 = mask_hexa(im_size[0]-self.border_size,im_size[1]-self.border_size,(im_size[1]-self.border_size)//2)
        cache = {}
        for l in self.struct_map:
            if not l.active:
                continue
            if str(l.bg_rgb)+str(l.bd_rgb) in cache:
                bg = cache[str(l.bg_rgb)+str(l.bd_rgb)]
            else:
                bg = Image.new('RGBA', im_size,(*l.bg_rgb,0))
                bg.paste(Image.new('RGBA', im_size,l.bd_rgb),(0,0),mask1)
                bg.paste(Image.new('RGBA', [im_size[0]-self.border_size,im_size[1]-self.border_size],(*l.bg_rgb,255)),
                         (self.border_size//2,self.border_size//2),
                         mask2)
                cache[str(l.bg_rgb)+str(l.bd_rgb)] = bg
            
            self.carte.paste(bg,self.pos(l.x,l.y),bg)
    
    def add_fonds(self):
        """ajoute les images de fond"""
        im_size = (2*int(self.r*self._cos),2*self.r)
        mask1 = mask_hexa(*im_size,(im_size[1]-self.border_size)//2)
        
        temp = Image.new('RGBA', self.carte.size,(0,0,0,0))
        
        for l in self.struct_map:
            if not l.active or l.bg not in self.bgs:
                continue
            im = self.bgs.get(l.bg)
            temp.paste(im,self.pos(l.x,l.y),mask1)

        self.carte.alpha_composite(temp)
    
    def add_images(self):
        """ajoute les images"""
        
        im_size = (2*int(self.r*self._cos),2*self.r)
        
        conf_size = {
            3:[int(self.r/2),int(-self.r*3/4),1.5],
            4:[0,-int(self.r),1.6],
            7:[-im_size[0]+im_size[0]//3,-int(im_size[1]/2.5),2.5],
        }
        
        temp = Image.new('RGBA', self.carte.size,(0,0,0,0))
        for l in self.struct_map:
            if not l.active>1 or l.obj not in self.objs:
                continue
            
            im = self.objs.get(l.obj)
            ix,iy = im.size
            xh,yh = self.pos(l.x,l.y)
            off_x, off_y, coef = conf_size.get(l.obj_size,[0,0,1])

            if coef > 1:
                x,y = im.size
                im = im.resize((int(coef*x),int(coef*y)))
            
            temp.paste(im,(xh+off_x+(im_size[0]-ix)//2,yh+off_y+(im_size[1]-iy)//2),im)

        self.carte.alpha_composite(temp)

    def add_icons(self):
        """ajoute les icons"""
    
        im_size = (2*int(self.r*self._cos),2*self.r)
        
        conf_size = {
            1:[im_size[0]-im_size[0]//5-self.border_size*2,im_size[0]//4+self.border_size,5],
            3:[im_size[0]-im_size[0]//6,-im_size[0]//2,3],
            4:[im_size[0]-im_size[0]//3,-im_size[0]//2-im_size[0]//6,3],
            7:[im_size[0]-im_size[0]//6,-im_size[0]//2,3],
        }

        mask1 = mask_hexa(*im_size,(im_size[1]-self.border_size)//2)
        temp = Image.new('RGBA', self.carte.size,(0,0,0,0))
        for l in self.struct_map:
            if not l.active>2 or l.icon not in self.icons:
                continue

            im = self.icons.get(l.icon)

            off_x, off_y, coef = conf_size.get(l.obj_size or 1,[0,0,5])

            im = im.copy().resize((im.size[0]//coef,im.size[1]//coef))

            
            ix,iy = im.size
            xh,yh = self.pos(l.x,l.y)
            
            x,y = (xh+off_x+(im.size[0]-ix)//2,yh+off_y+(im.size[1]-iy)//2)
            
            temp.paste(Image.new('RGBA', im.size,(0,0,0,255)),(x+3,y+3),im)
            temp.paste(im,(x,y),im)
            
        self.carte.alpha_composite(temp,)

    def pos(self,x,y):
        x,y = x-self.offset[0],y-self.offset[1]
        out_x,out_y = 0,0
        if x:
            out_x += x*2*int(self.r*self._cos)
        if y:
            out_y += y*int(1.5*self.r)
        if y%2:
            out_x += int(self.r*self._cos)
        return out_x,out_y

    def make_carte(self):
        x,y = self.size
        out_x,out_y = 0,0
        out_y = int(y*1.5*self.r+self.r/2)
        out_x = x*int(2*self.r*self._cos)

        return Image.new('RGBA', (out_x,out_y),(255,255,255,0))
    
    def append_hexa(self,im_name,coord,bg_rgb=(0,0,0),bd_rgb=(0,0,0,255),border_size=10,obj=None,icon=None):
        im = self.make_border_hexa(im_name,border_size=border_size,bg_rgb=bg_rgb,bd_rgb=bd_rgb,obj_name=obj,icon_name=icon)
        self.carte.paste(im,self.pos(*coord),im)
    
    def aff_coord(self,color=(255,255,100)):
        carte = self.carte.copy()
        obj = ImageDraw.Draw(carte)
        fnt = ImageFont.truetype('./aAhaWow.ttf', 100)
        for l in self.struct_map:
            pos_x,pos_y = self.pos(l.x,l.y)
            obj.text((pos_x+3+int(self.r*self._cos)//2,pos_y+5+self.r//2), "%i/%i" % (l.x,l.y), font=fnt,fill=(0,0,0))
            obj.text((pos_x+int(self.r*self._cos)//2,pos_y+self.r//2), "%i/%i" % (l.x,l.y), font=fnt,fill=color)

        return carte

    def generer(self):
        self.make_bg()
        self.add_fonds()
        self.add_images()
        self.add_icons()

    def reset(self):
        self.carte = self.make_carte()
